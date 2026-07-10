from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.utils.exceptions import InvalidFileException

HEADERS = ("부서", "이름", "사번", "전화번호")
MAX_ROWS = 500
PHONE_RE = re.compile(r"^010\d{8}$")


class BulkFileError(ValueError):
    """The uploaded file itself cannot be previewed."""


@dataclass(frozen=True)
class RawEmployeeRow:
    row: int
    department: str | None
    name: str
    employee_no: str | None
    phone: str
    auto_generated: bool = False


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def normalize_phone(value: object) -> str:
    return re.sub(r"[\s-]", "", clean_cell(value))


def company_prefix(company_id: str) -> str:
    prefix = "".join(ch for ch in company_id.upper() if ch.isalnum())[:3]
    return prefix or "EMP"


def allocate_employee_no(company_id: str, used: set[str]) -> str:
    prefix = company_prefix(company_id)
    for sequence in range(1, 10000):
        candidate = f"{prefix}-{sequence:04d}"
        if candidate not in used:
            used.add(candidate)
            return candidate
    raise BulkFileError("자동 채번 가능한 사번을 모두 사용했어요")


def _nonempty_rows(rows: Iterable[Sequence[object]]) -> list[tuple[int, list[str]]]:
    result: list[tuple[int, list[str]]] = []
    for row_number, values in enumerate(rows, start=2):
        cells = [clean_cell(value) for value in list(values)[:4]]
        cells += [""] * (4 - len(cells))
        if any(cells):
            result.append((row_number, cells))
    return result


def read_employee_file(filename: str, content: bytes) -> list[RawEmployeeRow]:
    suffix = Path(filename or "").suffix.lower()
    if suffix not in {".xlsx", ".csv"}:
        raise BulkFileError(".xlsx 또는 .csv 파일만 업로드할 수 있어요")
    if not content:
        raise BulkFileError("파일이 비어 있어요")

    try:
        if suffix == ".csv":
            text = content.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            header = next(reader, [])
            data_rows = _nonempty_rows(reader)
        else:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            if sheet is None:
                raise BulkFileError("엑셀 워크시트를 찾을 수 없어요")
            iterator = sheet.iter_rows(values_only=True)
            header = list(next(iterator, ()))
            data_rows = _nonempty_rows(iterator)
            workbook.close()
    except (UnicodeDecodeError, csv.Error, ValueError, OSError, BadZipFile, KeyError, ParseError, InvalidFileException) as exc:
        raise BulkFileError("파일을 읽을 수 없어요. 양식을 다운로드해서 사용해주세요") from exc

    if tuple(clean_cell(value) for value in header) != HEADERS:
        raise BulkFileError("헤더와 순서가 올바르지 않아요. 양식을 다운로드해서 사용해주세요")
    if len(data_rows) > MAX_ROWS:
        raise BulkFileError("데이터 행은 최대 500행까지 업로드할 수 있어요")

    return [
        RawEmployeeRow(
            row=row_number,
            department=cells[0] or None,
            name=cells[1],
            employee_no=cells[2] or None,
            phone=normalize_phone(cells[3]),
        )
        for row_number, cells in data_rows
    ]


def validate_rows(
    rows: Sequence[RawEmployeeRow],
    *,
    company_id: str,
    existing_phones: Iterable[str] = (),
    existing_employee_nos: Iterable[str] = (),
) -> dict[str, list[dict]]:
    if len(rows) > MAX_ROWS:
        raise BulkFileError("데이터 행은 최대 500행까지 업로드할 수 있어요")

    db_phones = {normalize_phone(phone) for phone in existing_phones if phone}
    db_employee_nos = {clean_cell(number) for number in existing_employee_nos if number}
    used_employee_nos = set(db_employee_nos)
    seen_phones: set[str] = set()
    valid: list[dict] = []
    errors: list[dict] = []

    for raw in rows:
        name = clean_cell(raw.name)
        department = clean_cell(raw.department) or None
        phone = normalize_phone(raw.phone)
        manual_no = None if raw.auto_generated else (clean_cell(raw.employee_no) or None)
        reason = None
        if not name:
            reason = "이름 누락"
        elif not PHONE_RE.fullmatch(phone):
            reason = "전화번호 형식 오류"
        elif phone in seen_phones:
            reason = "파일 내 중복"
        elif phone in db_phones:
            reason = "이미 등록된 번호"
        elif manual_no and manual_no in db_employee_nos:
            reason = "이미 등록된 사번"
        elif manual_no and manual_no in used_employee_nos:
            reason = "파일 내 사번 중복"

        if PHONE_RE.fullmatch(phone):
            seen_phones.add(phone)
        if reason:
            errors.append({
                "row": raw.row,
                "reason": reason,
                "department": department,
                "name": name,
                "employee_no": manual_no,
                "phone": phone,
            })
            continue

        if manual_no:
            employee_no = manual_no
            used_employee_nos.add(employee_no)
            auto_generated = False
        else:
            employee_no = allocate_employee_no(company_id, used_employee_nos)
            auto_generated = True
        valid.append({
            "row": raw.row,
            "department": department,
            "name": name,
            "employee_no": employee_no,
            "phone": phone,
            "auto_generated": auto_generated,
        })
    return {"valid": valid, "errors": errors}


def build_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError("template worksheet unavailable")
    sheet.title = "직원 일괄등록"
    sheet.append(list(HEADERS))
    sheet.append(["홍보팀", "홍길동", "E001", "01012345678"])
    sheet["A1"].comment = Comment("사번/부서는 비워두셔도 됩니다", "greeneatGo")
    sheet.freeze_panes = "A2"
    widths = (34, 18, 18, 20)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
