from __future__ import annotations

import base64
import html
import io
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_HEADER_FILL = PatternFill("solid", fgColor="DDEEDB")
_TITLE_FILL = PatternFill("solid", fgColor="F6F0DF")
_KST = ZoneInfo("Asia/Seoul")


def _text(value: Any, fallback: str = "-") -> str:
    if value is None:
        return fallback
    rendered = str(value).strip()
    return rendered or fallback


def _amount(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _xlsx_text(value: Any, fallback: str = "-") -> str:
    rendered = _text(value, fallback)
    return f"'{rendered}" if rendered.startswith(("=", "+", "-", "@")) else rendered


def _transaction_date_time(value: Any) -> tuple[Any, Any]:
    rendered = _text(value, "")
    if not rendered:
        return "-", "-"
    try:
        parsed = datetime.fromisoformat(rendered.replace("Z", "+00:00"))
        local = parsed.astimezone(_KST) if parsed.tzinfo else parsed.replace(tzinfo=_KST)
        return local.date(), local.time().replace(tzinfo=None, microsecond=0)
    except ValueError:
        return rendered, "-"


def _original_invoice(row: dict[str, Any]) -> dict[str, Any]:
    return next(
        (
            invoice
            for invoice in row.get("tax_invoices") or []
            if isinstance(invoice, dict) and invoice.get("document_type") == "original"
        ),
        {},
    )


def _style_sheet(sheet, *, freeze: str | None = None) -> None:
    sheet.sheet_view.showGridLines = False
    if freeze:
        sheet.freeze_panes = freeze
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(
                name="Arial",
                size=cell.font.sz or 10,
                bold=cell.font.bold,
            )
            cell.alignment = Alignment(vertical="center")
    for column in range(1, sheet.max_column + 1):
        values = [len(str(sheet.cell(row=row, column=column).value or "")) for row in range(1, sheet.max_row + 1)]
        sheet.column_dimensions[get_column_letter(column)].width = min(max(max(values, default=8) + 3, 11), 34)


def _workbook_bytes(workbook: Workbook) -> bytes:
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_settlement_xlsx(detail: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "매출 정산"

    supplier = detail.get("supplier_information") or {}
    recipient = detail.get("business_information") or {}
    summary_rows = [
        ["매출 정산서"],
        ["정산 기간", _text(detail.get("period_from")), "~", _text(detail.get("period_to"))],
        ["공급자", _xlsx_text(supplier.get("name")), "사업자번호", _xlsx_text(supplier.get("biz_reg_no"))],
        ["공급받는자", _xlsx_text(recipient.get("name")), "사업자번호", _xlsx_text(recipient.get("biz_reg_no"))],
        ["공급가액", _amount(detail.get("supply_amount")), "부가세", _amount(detail.get("vat_amount")), "합계", _amount(detail.get("total_amount"))],
        ["정산 상태", _text(detail.get("settlement_status")), "세금계산서 상태", _text(detail.get("tax_invoice_status")), "입금 상태", _text(detail.get("payment_status"))],
        [],
    ]
    for row in summary_rows:
        sheet.append(row)

    headers = ["거래 날짜", "거래 시간", "이름", "부서", "사번", "구분", "내역", "공급가액", "부가세", "합계", "거래번호"]
    sheet.append(headers)
    header_row = sheet.max_row
    for transaction in detail.get("transactions") or []:
        kind = transaction.get("kind")
        kind_label = "장부" if kind == "spend" else "환불" if kind == "refund" else "취소"
        if kind == "spend" and transaction.get("pay_type") == "subsidized":
            kind_label = "보조금"
        transaction_date, transaction_time = _transaction_date_time(transaction.get("created_at"))
        sheet.append([
            transaction_date,
            transaction_time,
            _xlsx_text(transaction.get("employee_name")),
            _xlsx_text(transaction.get("department")),
            _xlsx_text(transaction.get("employee_no")),
            kind_label,
            _xlsx_text(transaction.get("item")),
            _amount(transaction.get("supply_amount")),
            _amount(transaction.get("vat_amount")),
            _amount(transaction.get("total_amount")),
            _xlsx_text(transaction.get("tx_code")),
        ])

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"].font = Font(name="Arial", size=16, bold=True)
    sheet["A1"].fill = _TITLE_FILL
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for cell in sheet[header_row]:
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in range(2, 7):
        for column in (2, 4, 6):
            sheet.cell(row=row, column=column).number_format = "#,##0"
    for row in range(header_row + 1, sheet.max_row + 1):
        sheet.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        sheet.cell(row=row, column=2).number_format = "hh:mm:ss"
        for column in (8, 9, 10):
            sheet.cell(row=row, column=column).number_format = "#,##0;[Red]-#,##0;0"
    sheet.auto_filter.ref = f"A{header_row}:K{sheet.max_row}"
    _style_sheet(sheet, freeze=f"A{header_row + 1}")
    return _workbook_bytes(workbook)


def vat_reference_rows(settlements: Iterable[dict[str, Any]], year: str) -> list[list[Any]]:
    rows: list[list[Any]] = []
    valid_statuses = {"issued", "nts_sending", "nts_accepted"}
    for settlement in settlements:
        if settlement.get("is_demo") or settlement.get("tax_invoice_status") not in valid_statuses:
            continue
        invoice = _original_invoice(settlement)
        write_date = _text(invoice.get("write_date"), "")
        if not write_date.startswith(year):
            continue
        recipient = invoice.get("recipient_snapshot") or settlement.get("business_information") or {}
        rows.append([
            "매출",
            f"{_text(settlement.get('period_ym'))} 식대(월합계)",
            "세금계산서",
            _xlsx_text(recipient.get("name")),
            _xlsx_text(recipient.get("biz_reg_no") or recipient.get("registration_number") or recipient.get("corp_num")),
            _amount(settlement.get("supply_amount")),
            _amount(settlement.get("vat_amount")),
            _amount(settlement.get("total_amount")),
            write_date,
            _xlsx_text(invoice.get("nts_confirm_num"), "미수신"),
            _xlsx_text(invoice.get("issued_at"), "미확정"),
        ])
    return rows


def build_vat_reference_xlsx(settlements: Iterable[dict[str, Any]], year: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "부가가치세 신고 참고자료"
    headers = ["증빙 구분", "품목명", "증빙 유형", "공급받는자", "사업자번호", "공급가액", "부가세", "총액", "작성일자", "승인번호", "승인일자"]
    sheet.append([f"{year}년 부가가치세 신고 참고자료"])
    sheet.append(["참고용 자료이며 최종 신고 전 세무 담당자와 확인해 주세요."])
    sheet.append([])
    sheet.append(headers)
    for row in vat_reference_rows(settlements, year):
        sheet.append(row)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sheet["A1"].font = Font(name="Arial", size=16, bold=True)
    sheet["A1"].fill = _TITLE_FILL
    sheet["A1"].alignment = Alignment(horizontal="center")
    for cell in sheet[4]:
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in range(5, sheet.max_row + 1):
        for column in (6, 7, 8):
            sheet.cell(row=row, column=column).number_format = "#,##0;[Red]-#,##0;0"
    sheet.auto_filter.ref = f"A4:K{sheet.max_row}"
    _style_sheet(sheet, freeze="A5")
    return _workbook_bytes(workbook)


def build_settlement_html(detail: dict[str, Any]) -> str:
    supplier = detail.get("supplier_information") or {}
    recipient = detail.get("business_information") or {}
    transactions = detail.get("transactions") or []

    def e(value: Any) -> str:
        return html.escape(_text(value))

    def transaction_row(row: dict[str, Any]) -> str:
        transaction_date, transaction_time = _transaction_date_time(row.get("created_at"))
        return (
            "<tr>"
            f"<td>{e(transaction_date)}</td><td>{e(transaction_time)}</td><td>{e(row.get('employee_name'))}</td>"
            f"<td>{e(row.get('department'))}</td><td>{e(row.get('employee_no'))}</td>"
            f"<td>{e(row.get('kind'))}</td><td>{e(row.get('item'))}</td>"
            f"<td class='money'>{_amount(row.get('supply_amount')):,}</td>"
            f"<td class='money'>{_amount(row.get('vat_amount')):,}</td>"
            f"<td class='money'>{_amount(row.get('total_amount')):,}</td>"
            f"<td>{e(row.get('tx_code'))}</td></tr>"
        )

    transaction_rows = "".join(transaction_row(row) for row in transactions) or (
        "<tr><td colspan='11' class='empty'>해당 정산 기간의 사용 내역이 없습니다.</td></tr>"
    )
    return f"""<!doctype html><html lang='ko'><head><meta charset='utf-8'><title>매출 정산서</title><style>
@page {{ size:A4 landscape; margin:16mm; }}
body {{ font-family:'Settlement Korean','Malgun Gothic',sans-serif; color:#17351f; margin:24px; font-size:12px; }}
h1 {{ margin:0 0 18px; font-size:24px; }} .summary {{ display:grid; grid-template-columns:repeat(3,1fr); gap:10px; margin-bottom:18px; }}
.summary div {{ background:#f7f2e5; border:1px solid #d9ddce; border-radius:8px; padding:10px; }} .summary span {{ display:block; color:#647066; font-size:10px; }}
table {{ width:100%; border-collapse:collapse; }} th,td {{ border-bottom:1px solid #d9ddce; padding:7px; text-align:left; }} th {{ background:#dfeedb; }} .money {{ text-align:right; }} .empty {{ text-align:center; padding:24px; }}
</style></head><body><h1>매출 정산서</h1><section class='summary'>
<div><span>정산 기간</span><strong>{e(detail.get('period_from'))} ~ {e(detail.get('period_to'))}</strong></div>
<div><span>공급자</span><strong>{e(supplier.get('name'))}</strong><br>{e(supplier.get('biz_reg_no'))}</div>
<div><span>공급받는자</span><strong>{e(recipient.get('name'))}</strong><br>{e(recipient.get('biz_reg_no'))}</div>
<div><span>공급가액</span><strong>{_amount(detail.get('supply_amount')):,}원</strong></div>
<div><span>부가세</span><strong>{_amount(detail.get('vat_amount')):,}원</strong></div>
<div><span>합계</span><strong>{_amount(detail.get('total_amount')):,}원</strong></div></section>
<table><thead><tr><th>거래 날짜</th><th>거래 시간</th><th>이름</th><th>부서</th><th>사번</th><th>구분</th><th>내역</th><th>공급가액</th><th>부가세</th><th>합계</th><th>거래번호</th></tr></thead><tbody>{transaction_rows}</tbody></table></body></html>"""


def build_settlement_pdf(detail: dict[str, Any]) -> bytes:
    import importlib

    font_path = Path(__file__).resolve().parents[1] / "assets" / "fonts" / "NanumGothic-Regular.ttf"
    font_data = base64.b64encode(font_path.read_bytes()).decode("ascii")
    document = build_settlement_html(detail).replace(
        "</style>",
        f"@font-face {{ font-family:'Settlement Korean'; src:url(data:font/ttf;base64,{font_data}) format('truetype'); }}</style>",
        1,
    )
    return importlib.import_module("weasyprint").HTML(string=document).write_pdf()


XLSX_MEDIA_TYPE = _XLSX_MEDIA_TYPE
