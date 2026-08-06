import io
from datetime import datetime, time

from openpyxl import load_workbook

from app.services.settlement_documents import (
    build_settlement_html,
    build_settlement_pdf,
    build_settlement_xlsx,
    build_vat_reference_xlsx,
    vat_reference_rows,
)


def detail():
    return {
        "period_ym": "2026-07",
        "period_from": "2026-07-01",
        "period_to": "2026-07-31",
        "supply_amount": 10000,
        "vat_amount": 1000,
        "total_amount": 11000,
        "settlement_status": "confirmed",
        "tax_invoice_status": "issued",
        "payment_status": "unpaid",
        "supplier_information": {"name": "돈토식당", "biz_reg_no": "123-45-67890"},
        "business_information": {"name": "그린회사", "biz_reg_no": "111-22-33333"},
        "transactions": [
            {
                "created_at": "2026-07-03T03:34:00Z",
                "employee_name": "=HYPERLINK(\"https://bad.example\")",
                "department": "영업팀",
                "employee_no": "A-1",
                "kind": "spend",
                "pay_type": "ledger",
                "item": "중식",
                "supply_amount": 10000,
                "vat_amount": 1000,
                "total_amount": 11000,
                "tx_code": "TX-1",
            }
        ],
        "tax_invoices": [
            {
                "document_type": "original",
                "write_date": "2026-07-31",
                "recipient_snapshot": {"name": "그린회사", "biz_reg_no": "111-22-33333"},
                "nts_confirm_num": "NTS-1",
                "issued_at": "2026-08-01T09:00:00+09:00",
            }
        ],
    }


def test_settlement_xlsx_contains_summary_transactions_and_no_formulas():
    workbook = load_workbook(io.BytesIO(build_settlement_xlsx(detail())), data_only=False)
    sheet = workbook["매출 정산"]

    assert sheet["A1"].value == "매출 정산서"
    assert sheet["B5"].value == 10000
    assert sheet["F5"].value == 11000
    assert [cell.value for cell in sheet[8]][:6] == ["거래 날짜", "거래 시간", "이름", "부서", "사번", "구분"]
    assert sheet["A9"].value == datetime(2026, 7, 3)
    assert sheet["B9"].value == time(12, 34)
    assert sheet["C9"].value.startswith("'=")
    assert sheet["D9"].value == "영업팀"
    assert sheet["E9"].value == "A-1"
    assert sheet["H9"].value == 10000
    assert sheet["A9"].number_format == "yyyy-mm-dd"
    assert sheet["B9"].number_format == "hh:mm:ss"
    assert sheet.auto_filter.ref == "A8:K9"
    assert sheet.freeze_panes == "A9"


def test_vat_reference_filters_to_issued_documents_in_selected_year():
    old = {**detail(), "period_ym": "2025-12", "tax_invoices": [{**detail()["tax_invoices"][0], "write_date": "2025-12-31"}]}
    pending = {**detail(), "tax_invoice_status": "requested"}

    rows = vat_reference_rows([detail(), old, pending], "2026")

    assert len(rows) == 1
    assert rows[0][0:3] == ["매출", "2026-07 식대(월합계)", "세금계산서"]
    assert rows[0][5:8] == [10000, 1000, 11000]
    workbook = load_workbook(io.BytesIO(build_vat_reference_xlsx([detail(), old, pending], "2026")))
    sheet = workbook["부가가치세 신고 참고자료"]
    assert sheet["A1"].value == "2026년 부가가치세 신고 참고자료"
    assert sheet.max_row == 5
    assert sheet["F5"].value == 10000


def test_settlement_view_and_pdf_embed_korean_content_and_font():
    document = build_settlement_html(detail())
    assert "매출 정산서" in document
    assert "돈토식당" in document
    assert "그린회사" in document
    assert "=HYPERLINK" in document
    assert "<th>거래 날짜</th><th>거래 시간</th>" in document
    assert "<th>부서</th><th>사번</th>" in document
    assert "<td>2026-07-03</td><td>12:34:00</td>" in document
    assert "영업팀 / A-1" not in document
    assert "<td>장부</td>" in document
    assert "<td>spend</td>" not in document
    assert "<th class='money'>공급가액</th><th class='money'>부가세</th><th class='money'>합계</th>" in document
    assert ".money { width:100px; min-width:100px; max-width:100px;" in document
    assert "text-align:right; white-space:nowrap; font-variant-numeric:tabular-nums;" in document
    assert "<th class='transaction-code'>거래번호</th>" in document
    assert ".transaction-code { text-align:left; }" in document

    pdf = build_settlement_pdf(detail())
    assert pdf.startswith(b"%PDF-")
    assert len(pdf) > 10000
